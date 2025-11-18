# =========================
# Configuration utilisateur
# =========================
from pathlib import Path

# Nom des fichiers d'entrée (adapter si besoin)
TAUX_FILE = "Equant Corporate  rates_October_2025.xlsx"   # Fichier de taux
EXPENSES_FILE = "expenses2393.xlsx"                       # Fichier des dépenses

# Feuille à lire pour les dépenses
EXPENSES_SHEET = "Tab Expenses"

# Liste des jobs (cohérente avec la demande)
# - type="include": garder les lignes dont COST CENTER contient le code
# - type="exclude": exclure les lignes dont COST CENTER contient l'un des codes listés
COST_CENTER_JOBS = [
    {"name": "MOBWE", "type": "include", "codes": ["MOBWE"], "sheets": ["Tab Expenses", "RefactExpenses"]},
    {"name": "BWDY8", "type": "include", "codes": ["BWDY8"], "sheets": ["Tab Expenses", "RefactExpenses"]},
    {"name": "BWDI3", "type": "exclude", "codes": ["MOBWE", "BWDY8"], "sheets": ["Tab Expenses", "MOBTJ", "RefactExpenses"]},
]

# Dossier de sortie (optionnel)
OUTPUT_DIR = Path(".")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# =========================
# Imports & utilitaires
# =========================
import os
import calendar
import unicodedata
from datetime import datetime
from dateutil.relativedelta import relativedelta
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import zipfile
import io

# 👉 Ajout Streamlit (seule vraie dépendance nouvelle)
import streamlit as st


def normalize_str(s):
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return ""
    txt = str(s).strip().lower()
    txt = unicodedata.normalize("NFKD", txt).encode("ASCII", "ignore").decode("utf-8")
    return txt


MOIS_FR_VARIANTS = {
    "jan": "janvier", "janv": "janvier", "janv.": "janvier", "janvier": "janvier",
    "fev": "fevrier", "fevr": "fevrier", "fevr.": "fevrier", "fevrier": "fevrier", "févr.": "fevrier", "février": "fevrier",
    "mar": "mars", "mars": "mars",
    "avr": "avril", "avr.": "avril", "avril": "avril",
    "mai": "mai",
    "juin": "juin",
    "juil": "juillet", "juil.": "juillet", "juillet": "juillet",
    "aout": "aout", "août": "aout", "aout.": "aout",
    "sept": "septembre", "sept.": "septembre", "sep": "septembre", "september": "septembre", "septembre": "septembre",
    "oct": "octobre", "oct.": "octobre", "octobre": "octobre",
    "nov": "novembre", "nov." : "novembre", "novembre": "novembre",
    "dec": "decembre", "dec.": "decembre", "decembre": "decembre", "déc.": "decembre", "décembre": "decembre",
}
MOIS_EN_TO_FR = {
    "january": "janvier", "february": "fevrier", "march": "mars", "april": "avril",
    "may": "mai", "june": "juin", "july": "juillet", "august": "aout",
    "september": "septembre", "october": "octobre", "november": "novembre", "december": "decembre"
}


def resolve_previous_month_labels(today=None):
    if today is None:
        today = datetime.today()
    mois_prec = today - relativedelta(months=1)
    en_full = mois_prec.strftime("%B").lower()
    fr_full = MOIS_EN_TO_FR[en_full]
    excel_abbr = mois_prec.strftime("%b").upper()
    mmYYYY = mois_prec.strftime("%m%Y")
    last_day = calendar.monthrange(mois_prec.year, mois_prec.month)[1]
    date_fin = datetime(mois_prec.year, mois_prec.month, last_day).strftime("%d%m%Y")
    return {"mois_prec": mois_prec, "fr_full": fr_full, "en_full": en_full, "excel_abbr": excel_abbr, "mmYYYY": mmYYYY, "date_fin": date_fin}


def parse_facturation_month(value):
    v = normalize_str(value)
    if not v:
        return ""
    parts = v.replace("-", " ").replace("/", " ").split()
    mois, annee = "", ""
    for token in parts:
        if not mois:
            mois = MOIS_FR_VARIANTS.get(token, "")
        if not annee and token.isdigit() and len(token) == 4:
            annee = token
    if mois and annee:
        return f"{mois} {annee}"
    return v


# =========================
# Lecture & traitement commun
# =========================
def read_rates_table(filepath: str) -> pd.DataFrame:
    df_taux = pd.read_excel(filepath, skiprows=7)
    df_taux_clean = df_taux.iloc[:, 2:].copy()
    for col in df_taux_clean.columns:
        if col.upper() == "ISO-CODE":
            continue
        df_taux_clean[col] = (
            df_taux_clean[col].astype(str)
            .str.replace(",", ".", regex=False)
            .apply(lambda x: float(x) if x.replace('.', '', 1).isdigit() else None)
        )
    if "ISO-CODE" in df_taux_clean.columns:
        df_taux_clean["ISO-CODE"] = df_taux_clean["ISO-CODE"].astype(str).str.upper()
    return df_taux_clean


def read_expenses(filepath: str, sheet_name: str) -> pd.DataFrame:
    # 📥 Lecture du fichier
    df = pd.read_excel(filepath, sheet_name=sheet_name)

    # 🧹 Suppression des colonnes inutiles
    colonnes_a_supprimer = [
        "EMPLOYEE N°", "REGIONAL HR", "DEPARTEMENT", "AMOUNT IN USD",
        "AMOUNT IN GBP", "VAT IN EURO", "VAT IN USD", "VAT IN GBP", "USD RATE", "GBP RATE",
        "EXPENSE CREATION DATE", "PAIEMENT DATE", "DETAIL", "PROFORMAT",
        "COMMENTS", "DEVISE EMPLOYEE", "COMMANDE INTERNE"
    ]
    df = df.drop(columns=colonnes_a_supprimer, errors="ignore")

    # 🧭 Insertion ou repositionnement des colonnes clés
    if "INVOICE N°" in df.columns:
        col_index_invoice = df.columns.get_loc("INVOICE N°")

        # 🔹 Si AMOUNT IN EURO existe déjà → on la retire puis on la réinsère proprement
        if "AMOUNT IN EURO" in df.columns:
            col_data = df.pop("AMOUNT IN EURO")
            df.insert(col_index_invoice + 1, "AMOUNT IN EURO", col_data)
        else:
            df.insert(col_index_invoice + 1, "AMOUNT IN EURO", None)

        # 🔹 Pour VAT et TOTAL AMOUNT VAT INC → insérer uniquement si elles n'existent pas déjà
        if "VAT" not in df.columns:
            df.insert(col_index_invoice + 2, "VAT", "")
        if "TOTAL AMOUNT VAT INC" not in df.columns:
            df.insert(col_index_invoice + 3, "TOTAL AMOUNT VAT INC", "")
    else:
        raise ValueError("Colonne 'INVOICE N°' introuvable.")

    return df


def filter_previous_month(df: pd.DataFrame, mois_label_fr: str) -> pd.DataFrame:
    if "FACTURATION MONTH" not in df.columns:
        raise ValueError("Colonne 'FACTURATION MONTH' manquante.")
    df_tmp = df.copy()
    df_tmp["_FACT_MONTH_CLEAN_"] = df_tmp["FACTURATION MONTH"].apply(parse_facturation_month)
    mois_target = parse_facturation_month(mois_label_fr)
    out = df_tmp[df_tmp["_FACT_MONTH_CLEAN_"] == mois_target].copy()
    return out.drop(columns=["_FACT_MONTH_CLEAN_"])


def attach_currency_rate(df_exp: pd.DataFrame, df_rates: pd.DataFrame, excel_abbr: str) -> pd.DataFrame:
    if "CURRENCY" not in df_exp.columns:
        raise ValueError("Colonne 'CURRENCY' manquante.")
    df = df_exp.copy()
    df["CURRENCY"] = df["CURRENCY"].astype(str).str.upper()
    if excel_abbr not in df_rates.columns:
        raise ValueError(f"Colonne de taux '{excel_abbr}' introuvable.")
    rates = {}
    if "ISO-CODE" in df_rates.columns:
        for _, row in df_rates.iterrows():
            rates[str(row["ISO-CODE"]).upper()] = row[excel_abbr]

    def map_rate(code):
        return 1.0 if code == "EUR" else rates.get(code, None)

    df["CURRENCY RATE"] = df["CURRENCY"].apply(map_rate)
    return df


# =========================
# Export – feuilles (Tab, MOBTJ, Refact)
# =========================
def export_with_refact(sheet_df: pd.DataFrame, outfile: Path, sheets: list, refact_period_ddmmyyyy: str):
    with pd.ExcelWriter(outfile, engine="openpyxl") as writer:
        if "Tab Expenses" in sheets:
            sheet_df.to_excel(writer, sheet_name="Tab Expenses", index=False)
        if "MOBTJ" in sheets:
            pd.DataFrame().to_excel(writer, sheet_name="MOBTJ", index=False)
        if "RefactExpenses" in sheets:
            colonnes_refact = ["COST CENTER", "N° facture", "Prestation", "NOM PRENOM", "PERIODE", "AMOUNT IN EURO"]
            df_refact = pd.DataFrame([{
                "COST CENTER": "",
                "N° facture": "",
                "Prestation": "Other expenses - Executive Relocations - autres charges",
                "NOM PRENOM": "",
                "PERIODE": refact_period_ddmmyyyy,
                "AMOUNT IN EURO": ""
            }], columns=colonnes_refact)
            df_refact.to_excel(writer, sheet_name="RefactExpenses", index=False)


def insert_eur_formula_inplace(excel_path: Path):
    wb = load_workbook(excel_path)
    if "Tab Expenses" not in wb.sheetnames:
        wb.save(excel_path)
        return
    ws = wb["Tab Expenses"]
    col_total = col_rate = col_euro = None
    for idx, cell in enumerate(ws[1], start=1):
        titre = str(cell.value).strip().upper() if cell.value is not None else ""
        if titre == "TOTAL AMOUNT IN CURRENCY":
            col_total = idx
        elif titre == "CURRENCY RATE":
            col_rate = idx
        elif titre == "AMOUNT IN EURO":
            col_euro = idx
    if None in (col_total, col_rate, col_euro):
        wb.save(excel_path)
        raise ValueError("Colonnes manquantes pour la formule EUR.")
    for row in range(2, ws.max_row + 1):
        v_total = ws.cell(row=row, column=col_total).value
        v_rate = ws.cell(row=row, column=col_rate).value
        if isinstance(v_total, (int, float)) and isinstance(v_rate, (int, float)) and v_rate != 0:
            lt = get_column_letter(col_total)
            lr = get_column_letter(col_rate)
            ws.cell(row=row, column=col_euro).value = f"=({lt}{row}/{lr}{row})"
        else:
            ws.cell(row=row, column=col_euro).value = None
    wb.save(excel_path)


# =========================
# Nettoyage noms & fichier de vérif
# =========================
def export_lines_to_review(df_month: pd.DataFrame, mmYYYY: str):
    def is_invalid_name(val):
        if val is None or (isinstance(val, float) and pd.isna(val)) or str(val).strip() == "":
            return True
        txt = normalize_str(val)
        return txt in {"vide", "missing", "assignee", "entity"}

    df = df_month.copy()
    df["_FN_CLEAN_"] = df["FIRST NAME"].apply(normalize_str) if "FIRST NAME" in df.columns else ""
    df["_LN_CLEAN_"] = df["LAST NAME"].apply(normalize_str) if "LAST NAME" in df.columns else ""
    mask_verif = (
        df["_FN_CLEAN_"].isin(["assignee", "entity"]) |
        df["_LN_CLEAN_"].isin(["missing", "entity"])
    )
    df_a_verifier = df[mask_verif].copy()
    for c in ["_FN_CLEAN_", "_LN_CLEAN_"]:
        if c in df_a_verifier.columns:
            df_a_verifier.drop(columns=[c], inplace=True)
    out_name = f"OROW_Expenses_Lignes_a_verifier_{mmYYYY}.xlsx"
    df_a_verifier.to_excel(out_name, index=False)
    print(f"✅ Fichier des lignes à vérifier exporté : {out_name} (lignes: {len(df_a_verifier)})")


def clean_names_and_assignment(df_month: pd.DataFrame) -> pd.DataFrame:
    def replace_invalid(val):
        if val is None or (isinstance(val, float) and pd.isna(val)) or str(val).strip() == "":
            return "DMI"
        txt = normalize_str(val)
        if txt in {"vide", "missing", "assignee", "entity"}:
            return "DMI"
        return val

    df = df_month.copy()
    if "FIRST NAME" in df.columns:
        df["FIRST NAME"] = df["FIRST NAME"].apply(replace_invalid).astype(str).str.upper()
    if "LAST NAME" in df.columns:
        df["LAST NAME"] = df["LAST NAME"].apply(replace_invalid).astype(str).str.upper()
    if "TYPE OF ASSIGNMENT" in df.columns:
        def norm_assign(v):
            if v is None:
                return "N/A"
            t = str(v).strip().lower()
            return "N/A" if t == "cost estimate" or t == "" else v
        df["TYPE OF ASSIGNMENT"] = df["TYPE OF ASSIGNMENT"].apply(norm_assign)
    return df


# =========================
# Contrôle & génération _CONTROLE
# =========================
def generate_control_file(base_month_df: pd.DataFrame, target_file: Path, delete_base: bool = True) -> Path:
    COL_NOM = "LAST NAME"
    COL_PRENOM = "FIRST NAME"
    COL_COST = "COST CENTER"

    # ===============================
    # 🧮 Règle 1 - Détection des conflits nom/prénom sur plusieurs Cost Center
    # ===============================
    for col in (COL_NOM, COL_PRENOM, COL_COST):
        if col not in base_month_df.columns:
            raise ValueError(f"Colonne manquante: {col}")

    cc_counts = (
        base_month_df
        .assign(__LN=lambda d: d[COL_NOM].astype(str).str.strip(),
                __FN=lambda d: d[COL_PRENOM].astype(str).str.strip())
        .groupby(["__LN", "__FN"])[COL_COST]
        .nunique()
    )
    conflicted = set(cc_counts[cc_counts > 1].index)

    # ===============================
    # 🟠 Règle 2 - BWDI3 : cost center = MOBWE ou BWDY8 ou ASB75
    # ===============================
    cc_orange = {"MOBWE", "BWDY8", "ASB75"}

    # Création de la version _CONTROLE
    ctrl_path = target_file.with_name(f"{target_file.stem}_CONTROLE.xlsx")
    from shutil import copyfile
    copyfile(target_file, ctrl_path)

    # ===============================
    # 🎨 Application des surlignages
    # ===============================
    wb2 = load_workbook(ctrl_path)
    if "Tab Expenses" not in wb2.sheetnames:
        wb2.save(ctrl_path)
        print(f"⚠️ Feuille 'Tab Expenses' absente dans {ctrl_path.name}, contrôle ignoré.")
        return ctrl_path

    ws = wb2["Tab Expenses"]

    # Styles
    fill_rouge = PatternFill(fill_type="solid", start_color="FFFF0000", end_color="FFFF0000")
    fill_orange = PatternFill(fill_type="solid", start_color="FFFFA500", end_color="FFFFA500")
    font_gras = Font(bold=True, color="000000")

    # Index des colonnes
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    idx_last = headers.index(COL_NOM) + 1
    idx_first = headers.index(COL_PRENOM) + 1
    idx_cost = headers.index(COL_COST) + 1

    count_conflict = 0
    count_specifique = 0

    for r in range(2, ws.max_row + 1):
        ln = str(ws.cell(row=r, column=idx_last).value or "").strip()
        fn = str(ws.cell(row=r, column=idx_first).value or "").strip()
        cc_val = str(ws.cell(row=r, column=idx_cost).value or "").strip().upper()

        # 🟥 Règle 1 : conflits nom/prénom
        if (ln, fn) in conflicted:
            count_conflict += 1
            for cell in ws[r]:
                cell.fill = fill_rouge
                cell.font = font_gras

        # 🟧 Règle 2 : seulement pour BWDI3
        if "bwdi3" in target_file.stem.lower():
            if cc_val in cc_orange:
                count_specifique += 1
                for cell in ws[r]:
                    cell.fill = fill_orange
                ws.cell(row=r, column=idx_cost).font = font_gras  # gras uniquement sur COST CENTER

    wb2.save(ctrl_path)
    print("Contrôle terminé ✅")
    print(f"- Lignes surlignées en rouge (Règle 1)      : {count_conflict}")
    print(f"- Lignes surlignées en orange (Règle 2)     : {count_specifique}")
    print(f"- Fichier final généré                      : {ctrl_path.name}")

    if delete_base:
        try:
            os.remove(target_file)
        except Exception:
            pass

    return ctrl_path


# =========================
# MAIN – Exécution complète (logique inchangée)
# =========================
def main():
    labels = resolve_previous_month_labels()
    mois_prec = labels["mois_prec"]
    fr_full = labels["fr_full"]
    excel_abbr = labels["excel_abbr"]
    mmYYYY = labels["mmYYYY"]
    date_fin = labels["date_fin"]
    print(f"Mois ciblé : {fr_full} {mois_prec.year} (abbr taux: {excel_abbr}, mmYYYY: {mmYYYY})")

    df_rates = read_rates_table(TAUX_FILE)
    df_exp = read_expenses(EXPENSES_FILE, EXPENSES_SHEET)

    df_exp_filtered = filter_previous_month(df_exp, f"{fr_full} {mois_prec.year}")
    df_exp_filtered = attach_currency_rate(df_exp_filtered, df_rates, excel_abbr)

    # Export des lignes à vérifier (avant corrections)
    export_lines_to_review(df_exp_filtered, mmYYYY)

    # Nettoyage noms + assignment + MAJUSCULES
    df_exp_filtered = clean_names_and_assignment(df_exp_filtered)

    # Garder dataset complet du mois pour contrôle inter-cost center
    base_month_df_for_control = df_exp_filtered.copy()

    for job in COST_CENTER_JOBS:
        name = job["name"]
        jtype = job["type"]
        codes = [str(c).upper() for c in job["codes"]]
        sheets = job["sheets"]
        df_job = df_exp_filtered.copy()
        if jtype == "include":
            mask = False
            for code in codes:
                mask = (mask | df_job["COST CENTER"].astype(str).str.upper().str.contains(code, na=False))
            df_job = df_job[mask].copy()
        elif jtype == "exclude":
            mask = False
            for code in codes:
                mask = (mask | df_job["COST CENTER"].astype(str).str.upper().str.contains(code, na=False))
            df_job = df_job[~mask].copy()
        else:
            raise ValueError(f"Type de job inconnu: {jtype}")

        # Retirer colonnes techniques si présentes
        cols_to_drop = [c for c in df_job.columns if c.startswith('_') and c.endswith('_')]
        if cols_to_drop:
            df_job.drop(columns=cols_to_drop, inplace=True, errors="ignore")

        # Export base (sans contrôle) puis insertion formules puis version _CONTROLE seule
        outfile_base = OUTPUT_DIR / f"expenses2393_{name.lower()}_final.xlsx"
        export_with_refact(df_job, outfile_base, sheets, refact_period_ddmmyyyy=date_fin)
        insert_eur_formula_inplace(outfile_base)
        ctrl_file = generate_control_file(base_month_df_for_control, outfile_base, delete_base=True)
        print(f"✅ Sortie (contrôlée) : {ctrl_file}")

    print("Traitement terminé 🎉")


# =========================
# Application Streamlit
# =========================
def main_streamlit():
    st.title("📒 Pipeline OROW – Traitement, Contrôle et Génération")
    st.write(
        "Cette application exécute le même traitement que le notebook :\n"
        "- Filtre sur le mois de facturation\n"
        "- Jointure avec les taux de change\n"
        "- Nettoyage des noms / assignments\n"
        "- Génération des fichiers finalisés et contrôlés par cost center\n"
        "- Export du fichier des lignes à vérifier"
    )

    st.subheader("1️⃣ Chargement des fichiers d'entrée")
    with st.form("orow_form"):
        taux_file = st.file_uploader(
            "Fichier de taux de devise (ex : Equant Corporate rates_October_2025.xlsx)",
            type=["xlsx", "xls"],
            key="taux_file"
        )
        expenses_file = st.file_uploader(
            "Fichier des dépenses OROW (ex : expenses2393.xlsx)",
            type=["xlsx", "xls"],
            key="expenses_file"
        )
        run_btn = st.form_submit_button("Lancer le traitement")

    if run_btn:
        if not taux_file or not expenses_file:
            st.error("Merci de charger les deux fichiers (taux de devise ET expenses OROW).")
            return

        # Sauvegarde des fichiers uploadés avec les mêmes noms attendus par le pipeline
        taux_path = OUTPUT_DIR / TAUX_FILE
        expenses_path = OUTPUT_DIR / EXPENSES_FILE

        with open(taux_path, "wb") as f:
            f.write(taux_file.getbuffer())
        with open(expenses_path, "wb") as f:
            f.write(expenses_file.getbuffer())

        st.info("Fichiers chargés. Lancement du traitement…")

        try:
            with st.spinner("Traitement en cours…"):
                main()
        except Exception as e:
            st.error(f"Une erreur est survenue pendant le traitement : {e}")
            return

        st.success("Traitement terminé 🎉")

        # Recalcule des labels pour récupérer mmYYYY et nom du fichier de vérification
        labels = resolve_previous_month_labels()
        mmYYYY = labels["mmYYYY"]

        # Fichiers attendus (identiques à la version notebook)
        expected_outputs = [
            OUTPUT_DIR / "expenses2393_mobwe_final_CONTROLE.xlsx",
            OUTPUT_DIR / "expenses2393_bwdy8_final_CONTROLE.xlsx",
            OUTPUT_DIR / "expenses2393_bwdi3_final_CONTROLE.xlsx",
            OUTPUT_DIR / f"OROW_Expenses_Lignes_a_verifier_{mmYYYY}.xlsx",
        ]

        st.subheader("2️⃣ Téléchargement des fichiers générés")

        for out_path in expected_outputs:
            if out_path.exists():
                with open(out_path, "rb") as f:
                    st.download_button(
                        label=f"📥 Télécharger {out_path.name}",
                        data=f,
                        file_name=out_path.name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
            else:
                st.warning(f"Fichier non trouvé (vérifier les logs) : {out_path.name}")


# ================================
# Création d'un ZIP regroupant tous les fichiers générés
# ================================
zip_buffer = io.BytesIO()

with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
    for out_path in expected_outputs:
        if out_path.exists():
            zipf.write(out_path, arcname=out_path.name)

zip_buffer.seek(0)

st.subheader("📦 Télécharger tous les fichiers en un seul ZIP")

st.download_button(
    label="📥 Télécharger le pack complet (ZIP)",
    data=zip_buffer,
    file_name=f"OROW_Exports_{mmYYYY}.zip",
    mime="application/zip"
)




# 👉 En mode script Streamlit, c'est cette fonction qui est appelée
if __name__ == "__main__":
    main_streamlit()

