# =========================
# Configuration utilisateur
# =========================
from pathlib import Path

TAUX_FILE = "Equant Corporate  rates_October_2025.xlsx"   # Fichier de taux corporate
EXPENSES_FILE = "expenses2394.xlsx"                       # Fichier des dépenses OBS
EXPENSES_SHEET = "Tab Expenses"

# Jobs OBS
OBS_JOBS = [
    {"name": "mob", "job_type": "MOB", "sheets": ["Tab Expenses", "RefactExpenses"]},
    {"name": "vie", "job_type": "VIE", "sheets": ["Tab Expenses"]},
    {"name": "",    "job_type": "GENERAL", "sheets": ["Tab Expenses"]},
]

OUTPUT_DIR = Path(".")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


# =========================
# Imports & utilitaires
# =========================
import os
import calendar
import unicodedata
from datetime import datetime
from dateutil.relativedelta import relativedelta
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# 👉 Ajout pour l'application Streamlit
import streamlit as st


def normalize_str(s):
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return ""
    txt = str(s).strip().lower()
    txt = unicodedata.normalize("NFKD", txt).encode("ASCII", "ignore").decode("utf-8")
    return txt

MOIS_EN_TO_FR = {
    "january": "janvier", "february": "fevrier", "march": "mars", "april": "avril",
    "may": "mai", "june": "juin", "july": "juillet", "august": "aout",
    "september": "septembre", "october": "octobre", "november": "novembre", "december": "decembre"
}

def resolve_previous_month_labels(today=None):
    if today is None:
        today = datetime.today()
    mois_prec = today - relativedelta(months=1)
    en_full = mois_prec.strftime("%B").lower()
    fr_full = MOIS_EN_TO_FR[en_full]
    excel_abbr = mois_prec.strftime("%b").upper()
    mmYYYY = mois_prec.strftime("%m%Y")
    last_day = calendar.monthrange(mois_prec.year, mois_prec.month)[1]
    date_fin = datetime(mois_prec.year, mois_prec.month, last_day).strftime("%d%m%Y")
    return {"mois_prec": mois_prec, "fr_full": fr_full, "excel_abbr": excel_abbr, "mmYYYY": mmYYYY, "date_fin": date_fin}


# =========================
# Lecture & traitement commun
# =========================
def read_rates_table(filepath: str) -> pd.DataFrame:
    df_taux = pd.read_excel(filepath, skiprows=7)
    df_taux_clean = df_taux.iloc[:, 2:].copy()
    # Normaliser ISO-CODE
    if "ISO-CODE" in df_taux_clean.columns:
        df_taux_clean["ISO-CODE"] = df_taux_clean["ISO-CODE"].astype(str).str.upper()
    # Convertir valeurs
    for col in df_taux_clean.columns:
        if col.upper() == "ISO-CODE":
            continue
        df_taux_clean[col] = (
            df_taux_clean[col].astype(str)
            .str.replace(",", ".", regex=False)
            .apply(lambda x: float(x) if x.replace('.', '', 1).isdigit() else None)
        )
    return df_taux_clean

def read_expenses(filepath: str, sheet_name: str) -> pd.DataFrame:
    df = pd.read_excel(filepath, sheet_name=sheet_name)
    colonnes_a_supprimer = [
        "AMOUNT IN GBP", "VAT IN EURO", "VAT IN USD", "VAT IN GBP", "USD RATE", "GBP RATE",
        "EXPENSE CREATION DATE", "PAIEMENT DATE", "DETAIL", "PROFORMAT",
        "COMMENTS", "DEVISE EMPLOYEE", "COMMANDE INTERNE"
    ]
    df = df.drop(columns=colonnes_a_supprimer, errors="ignore")

    # Positionner colonnes autour de INVOICE N°
    if "INVOICE N°" not in df.columns:
        raise ValueError("Colonne 'INVOICE N°' introuvable.")
    idx_inv = df.columns.get_loc("INVOICE N°")

    # AMOUNT IN USD juste après INVOICE N°
    if "AMOUNT IN USD" in df.columns:
        col = df.pop("AMOUNT IN USD")
        df.insert(idx_inv + 1, "AMOUNT IN USD", col)
    else:
        df.insert(idx_inv + 1, "AMOUNT IN USD", None)

    # AMOUNT IN EURO après INVOICE N° (si présent le replacer)
    if "AMOUNT IN EURO" in df.columns:
        col = df.pop("AMOUNT IN EURO")
        df.insert(idx_inv + 2, "AMOUNT IN EURO", col)
    else:
        df.insert(idx_inv + 2, "AMOUNT IN EURO", None)

    # VAT et TOTAL AMOUNT VAT INC ensuite si absents
    if "VAT" not in df.columns:
        df.insert(idx_inv + 3, "VAT", "")
    if "TOTAL AMOUNT VAT INC" not in df.columns:
        df.insert(idx_inv + 4, "TOTAL AMOUNT VAT INC", "")

    # Repositionner EMPLOYEE N° avant NATIONALITY si nécessaire
    if "EMPLOYEE N°" in df.columns and "NATIONALITY" in df.columns:
        emp = df.pop("EMPLOYEE N°")
        idx_nat = df.columns.get_loc("NATIONALITY")
        df.insert(idx_nat, "EMPLOYEE N°", emp)
    return df

def filter_previous_month(df: pd.DataFrame, mois_label_fr: str) -> pd.DataFrame:
    if "FACTURATION MONTH" not in df.columns:
        raise ValueError("Colonne 'FACTURATION MONTH' manquante.")
    df_tmp = df.copy()
    df_tmp["_FACT_MONTH_CLEAN_"] = df_tmp["FACTURATION MONTH"].apply(lambda v: normalize_str(v))
    mois_target = normalize_str(mois_label_fr)
    out = df_tmp[df_tmp["_FACT_MONTH_CLEAN_"] == mois_target].copy()
    return out.drop(columns=["_FACT_MONTH_CLEAN_"])

def attach_currency_rate(df_exp: pd.DataFrame, df_rates: pd.DataFrame, excel_abbr: str) -> pd.DataFrame:
    df = df_exp.copy()
    df["CURRENCY"] = df["CURRENCY"].astype(str).str.upper()
    if excel_abbr not in df_rates.columns:
        raise ValueError(f"Colonne de taux '{excel_abbr}' introuvable dans le fichier taux.")
    # mapping ISO -> rate for month
    rates = {}
    if "ISO-CODE" in df_rates.columns:
        for _, row in df_rates.iterrows():
            rates[str(row["ISO-CODE"]).upper()] = row[excel_abbr]
    df["CURRENCY RATE"] = df["CURRENCY"].apply(lambda c: 1.0 if c == "EUR" else rates.get(c, None))
    return df


# =========================
# Export & insertion formules
# =========================
def export_with_optional_refact(sheet_df: pd.DataFrame, outfile: Path, sheets: list, refact_period_ddmmyyyy: str):
    with pd.ExcelWriter(outfile, engine="openpyxl") as writer:
        if "Tab Expenses" in sheets:
            sheet_df.to_excel(writer, sheet_name="Tab Expenses", index=False)
        if "RefactExpenses" in sheets:
            colonnes_refact = ["COST CENTER", "N° facture", "Prestation", "NOM PRENOM", "PERIODE", "AMOUNT IN EURO"]
            df_refact = pd.DataFrame([{
                "COST CENTER": "",
                "N° facture": "",
                "Prestation": "Other expenses - Executive Relocations - autres charges",
                "NOM PRENOM": "",
                "PERIODE": refact_period_ddmmyyyy,
                "AMOUNT IN EURO": ""
            }], columns=colonnes_refact)
            df_refact.to_excel(writer, sheet_name="RefactExpenses", index=False)

def insert_formulas_eur_usd_inplace(excel_path: Path, usd_rate: float | None):
    wb = load_workbook(excel_path)
    if "Tab Expenses" not in wb.sheetnames:
        wb.save(excel_path)
        return
    ws = wb["Tab Expenses"]
    # locate columns
    headers = [str(c.value).strip().upper() if c.value is not None else "" for c in ws[1]]
    def find_idx(name):
        return headers.index(name) + 1 if name in headers else None
    col_total = find_idx("TOTAL AMOUNT IN CURRENCY")
    col_rate  = find_idx("CURRENCY RATE")
    col_eur   = find_idx("AMOUNT IN EURO")
    col_usd   = find_idx("AMOUNT IN USD")
    if None in (col_total, col_rate, col_eur):
        wb.save(excel_path)
        raise ValueError("Colonnes manquantes pour formules EUR.")
    for row in range(2, ws.max_row + 1):
        v_total = ws.cell(row=row, column=col_total).value
        v_rate  = ws.cell(row=row, column=col_rate).value
        if isinstance(v_total, (int, float)) and isinstance(v_rate, (int, float)) and v_rate != 0:
            lt = get_column_letter(col_total)
            lr = get_column_letter(col_rate)
            ws.cell(row=row, column=col_eur).value = f"=({lt}{row}/{lr}{row})"
            if col_usd and usd_rate is not None:
                ws.cell(row=row, column=col_usd).value = f"=({lt}{row}/{lr}{row})*{usd_rate}"
        else:
            ws.cell(row=row, column=col_eur).value = None
            if col_usd:
                ws.cell(row=row, column=col_usd).value = None
    wb.save(excel_path)


# =========================
# Extraction des lignes à vérifier & nettoyage
# =========================
def export_lines_to_review(df_month: pd.DataFrame, mmYYYY: str):
    df = df_month.copy()
    df["_FN_"] = df["FIRST NAME"].apply(normalize_str) if "FIRST NAME" in df.columns else ""
    df["_LN_"] = df["LAST NAME"].apply(normalize_str) if "LAST NAME" in df.columns else ""
    mask = (
        df["_FN_"].isin(["assignee", "entity"]) |
        df["_LN_"].isin(["missing", "entity"]) 
    )
    df_verif = df[mask].copy()
    for c in ("_FN_", "_LN_"):
        if c in df_verif.columns:
            df_verif.drop(columns=[c], inplace=True)
    out_name = f"OBS_Expenses_Lignes_a_verifier_{mmYYYY}.xlsx"
    df_verif.to_excel(out_name, index=False)
    print(f"✅ Lignes à vérifier exportées : {out_name} ({len(df_verif)} lignes)")

def clean_names_and_assignment(df_month: pd.DataFrame) -> pd.DataFrame:
    def repl(val):
        if val is None or (isinstance(val, float) and pd.isna(val)) or str(val).strip()=="":
            return "DMI"
        t = normalize_str(val)
        return "DMI" if t in {"vide", "missing", "assignee", "entity"} else val

    df = df_month.copy()
    if "FIRST NAME" in df.columns:
        df["FIRST NAME"] = df["FIRST NAME"].apply(repl).astype(str).str.upper()
    if "LAST NAME" in df.columns:
        df["LAST NAME"] = df["LAST NAME"].apply(repl).astype(str).str.upper()
    if "TYPE OF ASSIGNMENT" in df.columns:
        def norm_assign(v):
            if v is None:
                return "N/A"
            t = str(v).strip().lower()
            return "N/A" if t == "cost estimate" or t == "" else v
        df["TYPE OF ASSIGNMENT"] = df["TYPE OF ASSIGNMENT"].apply(norm_assign)
    return df


# =========================
# Contrôle (_CONTROLE) – règles Rouge/Orange
# =========================
def generate_control_file_obs(base_month_df: pd.DataFrame, target_file: Path, job_type: str, delete_base: bool = True) -> Path:
    COL_NOM, COL_PRENOM, COL_COST = "LAST NAME", "FIRST NAME", "COST CENTER"
    # Règle 1 – Conflits nom/prénom
    for col in (COL_NOM, COL_PRENOM, COL_COST):
        if col not in base_month_df.columns:
            raise ValueError(f"Colonne manquante: {col}")
    cc_counts = (
        base_month_df
        .assign(__LN=lambda d: d[COL_NOM].astype(str).str.strip(),
                __FN=lambda d: d[COL_PRENOM].astype(str).str.strip())
        .groupby(["__LN", "__FN"])[COL_COST]
        .nunique()
    )
    conflicted = set(cc_counts[cc_counts > 1].index)

    ctrl_path = target_file.with_name(f"{target_file.stem}_CONTROLE.xlsx")
    from shutil import copyfile
    copyfile(target_file, ctrl_path)

    wb2 = load_workbook(ctrl_path)
    if "Tab Expenses" not in wb2.sheetnames:
        wb2.save(ctrl_path)
        print(f"⚠️ Feuille 'Tab Expenses' absente dans {ctrl_path.name}, contrôle ignoré.")
        return ctrl_path
    ws = wb2["Tab Expenses"]

    # Styles
    fill_rouge = PatternFill(fill_type="solid", start_color="FFFF0000", end_color="FFFF0000")
    fill_orange = PatternFill(fill_type="solid", start_color="FFFFA500", end_color="FFFFA500")
    font_gras = Font(bold=True, color="000000")

    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    idx_last = headers.index(COL_NOM) + 1
    idx_first = headers.index(COL_PRENOM) + 1
    idx_cost = headers.index(COL_COST) + 1

    def is_orange(cost_val: str) -> bool:
        cc = (cost_val or "").upper().strip()
        if job_type.upper() == "MOB":
            return not cc.startswith("MOB")
        if job_type.upper() == "VIE":
            return cc.startswith("MOB")
        if job_type.upper() == "GENERAL":
            return cc.startswith("MOB")
        return False

    red_cnt = orange_cnt = 0
    for r in range(2, ws.max_row + 1):
        ln = str(ws.cell(row=r, column=idx_last).value or "").strip()
        fn = str(ws.cell(row=r, column=idx_first).value or "").strip()
        cc_val = str(ws.cell(row=r, column=idx_cost).value or "").strip()

        # Règle 1 – Rouge prioritaire
        if (ln, fn) in conflicted:
            red_cnt += 1
            for cell in ws[r]:
                cell.fill = fill_rouge
                cell.font = font_gras
            continue  # priorité au rouge

        # Règle 2 – Orange
        if is_orange(cc_val):
            orange_cnt += 1
            for cell in ws[r]:
                cell.fill = fill_orange
            ws.cell(row=r, column=idx_cost).font = font_gras

    wb2.save(ctrl_path)
    print("Contrôle OBS terminé ✅")
    print(f"- Lignes rouges (conflits nom/prénom): {red_cnt}")
    print(f"- Lignes orange (spécifique {job_type}): {orange_cnt}")
    print(f"- Fichier final généré: {ctrl_path.name}")

    if delete_base:
        try:
            os.remove(target_file)
        except Exception:
            pass
    return ctrl_path


# =========================
# MAIN – Exécution OBS
# =========================
def main():
    labels = resolve_previous_month_labels()
    mois_prec = labels["mois_prec"]
    fr_full = labels["fr_full"]
    excel_abbr = labels["excel_abbr"]
    mmYYYY = labels["mmYYYY"]
    date_fin = labels["date_fin"]
    print(f"Mois ciblé : {fr_full} {mois_prec.year} (abbr taux: {excel_abbr}, mmYYYY: {mmYYYY})")

    df_rates = read_rates_table(TAUX_FILE)
    df_exp = read_expenses(EXPENSES_FILE, EXPENSES_SHEET)
    df_exp = filter_previous_month(df_exp, f"{fr_full} {mois_prec.year}")
    df_exp = attach_currency_rate(df_exp, df_rates, excel_abbr)

    # Export des lignes à vérifier (avant corrections noms)
    export_lines_to_review(df_exp, mmYYYY)

    # Nettoyage noms + assignment
    df_exp = clean_names_and_assignment(df_exp)
    base_month_df_for_control = df_exp.copy()

    # Récupérer taux USD du mois
    usd_rate = None
    try:
        usd_rate_series = df_rates.loc[df_rates["ISO-CODE"] == "USD", excel_abbr]
        if not usd_rate_series.empty and pd.notnull(usd_rate_series.iloc[0]):
            usd_rate = float(usd_rate_series.iloc[0])
    except Exception:
        usd_rate = None

    for job in OBS_JOBS:
        name = job["name"]
        job_type = job["job_type"]
        sheets = job["sheets"]
        df_job = df_exp.copy()

        if job_type == "MOB":
            mask = df_job["COST CENTER"].astype(str).str.upper().str.startswith("MOB", na=False)
            df_job = df_job[mask].copy()
        elif job_type == "VIE":
            df_job = df_job[df_job["TYPE OF ASSIGNMENT"].astype(str).str.upper().eq("VIE")].copy()
            # supprimer EMPLOYEE N° si présent
            if "EMPLOYEE N°" in df_job.columns:
                df_job = df_job.drop(columns=["EMPLOYEE N°"])
        elif job_type == "GENERAL":
            mask_mob = df_job["COST CENTER"].astype(str).str.upper().str.startswith("MOB", na=False)
            mask_vie = df_job["TYPE OF ASSIGNMENT"].astype(str).str.upper().eq("VIE")
            df_job = df_job[~(mask_mob | mask_vie)].copy()
        else:
            raise ValueError(f"Job type inconnu: {job_type}")

        # Drop colonnes techniques éventuelles
        cols_to_drop = [c for c in df_job.columns if c.startswith('_') and c.endswith('_')]
        if cols_to_drop:
            df_job.drop(columns=cols_to_drop, inplace=True, errors="ignore")

        # Export temp + formules
        outfile_base = OUTPUT_DIR / f"expenses2394_{(name.lower() + '_' if name else '')}final.xlsx"
        export_with_optional_refact(df_job, outfile_base, sheets, refact_period_ddmmyyyy=date_fin)
        insert_formulas_eur_usd_inplace(outfile_base, usd_rate)
        # Contrôle (génère _CONTROLE et supprime le base)
        ctrl_file = generate_control_file_obs(base_month_df_for_control, outfile_base, job_type=job_type, delete_base=True)
        print(f"✅ Sortie contrôlée: {ctrl_file}")

    print("Traitement OBS terminé 🎉")


# =========================
# APPLICATION STREAMLIT
# =========================
def main_streamlit():
    st.title("🧾 OBS – Traitement Automatisé des Expenses")
    st.write(
        "Cette application exécute le traitement automatisé des dépenses OBS : "
        "intégration des données, contrôles et génération des fichiers Excel finaux conformément aux règles métier."
    )

    st.markdown("""
    **Fonctionnalités principales :**
    - Import des fichiers de dépenses et des taux de change  
    - Sélection des données du mois de facturation ciblé  
    - Normalisation des informations salariés et affectations  
    - Classification automatique des lignes : **MOB**, **VIE**, **GÉNÉRAL**  
    - Application des taux et calculs EUR/USD  
    - Contrôles qualité selon les règles OBS  
    - Production des fichiers Excel finaux par segment  
    """)


    st.header("1️⃣ Importer vos fichiers")

    with st.form("obs_form"):
        taux_file = st.file_uploader(
            "Fichier de taux de devise (ex : Equant Corporate rates_October_2025.xlsx)",
            type=["xlsx", "xls"],
            key="taux_file",
        )
        expenses_file = st.file_uploader(
            "Fichier des dépenses OBS (ex : expenses2394.xlsx)",
            type=["xlsx", "xls"],
            key="expenses_file",
        )
        run_btn = st.form_submit_button("Lancer le traitement")

    if run_btn:
        if not taux_file or not expenses_file:
            st.error("Merci de charger les deux fichiers (taux de devise ET dépenses OBS).")
            return

        # Sauvegarde locale avec les mêmes noms que dans le notebook
        with open(TAUX_FILE, "wb") as f:
            f.write(taux_file.getbuffer())
        with open(EXPENSES_FILE, "wb") as f:
            f.write(expenses_file.getbuffer())

        st.info("Fichiers chargés. Lancement du traitement…")

        try:
            with st.spinner("Traitement OBS en cours…"):
                main()
        except Exception as e:
            st.error(f"Une erreur est survenue pendant le traitement : {e}")
            return

        st.success("Traitement terminé 🎉")

        # Récupérer mmYYYY pour le nom du fichier de vérif
        labels = resolve_previous_month_labels()
        mmYYYY = labels["mmYYYY"]

        # Fichiers attendus (identiques au notebook)
        expected_outputs = [
            OUTPUT_DIR / "expenses2394_mob_final_CONTROLE.xlsx",
            OUTPUT_DIR / "expenses2394_vie_final_CONTROLE.xlsx",
            OUTPUT_DIR / "expenses2394_final_CONTROLE.xlsx",
            OUTPUT_DIR / f"OBS_Expenses_Lignes_a_verifier_{mmYYYY}.xlsx",
        ]

        st.header("2️⃣ Télécharger les fichiers générés")

        for out_path in expected_outputs:
            if out_path.exists():
                with open(out_path, "rb") as f:
                    st.download_button(
                        label=f"📥 Télécharger {out_path.name}",
                        data=f,
                        file_name=out_path.name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
            else:
                st.warning(f"Fichier non trouvé (vérifier les logs) : {out_path.name}")


# Pour un script streamlit / module
if __name__ == "__main__":
    main_streamlit()

